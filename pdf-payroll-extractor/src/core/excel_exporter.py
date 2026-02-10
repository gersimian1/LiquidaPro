"""
Exportador de datos a Excel con formato profesional.
"""
from pathlib import Path
from typing import List, Dict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import logging

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exporta datos a archivos Excel con formato"""
    
    def export_to_excel(
        self,
        df: pd.DataFrame,
        output_path: str,
        totals: Dict[str, float],
        title: str = "Resumen de Liquidaciones"
    ) -> None:
        """
        Exporta DataFrame a Excel con formato profesional.
        
        Args:
            df: DataFrame con los datos
            output_path: Ruta del archivo de salida
            totals: Diccionario con totales por columna
            title: Título del reporte
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Exportar a Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Liquidaciones', index=False, startrow=2)
            
            # Obtener workbook y worksheet para formatear
            workbook = writer.book
            worksheet = writer.sheets['Liquidaciones']
            
            # Agregar título
            worksheet['A1'] = title
            worksheet['A1'].font = Font(bold=True, size=14, color='FFFFFF')
            worksheet['A1'].fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
            worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
            worksheet.merge_cells(f'A1:{self._get_column_letter(len(df.columns))}1')
            
            # Formatear encabezados
            header_row = 3
            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=header_row, column=col_num)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self._get_border()
            
            # Formatear datos
            for row_num in range(header_row + 1, header_row + len(df) + 1):
                for col_num, column in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.border = self._get_border()
                    
                    # Alinear nombres a la izquierda, números a la derecha
                    if column == 'Apellido y Nombre':
                        cell.alignment = Alignment(horizontal='left')
                    elif column == '% Aporte Jub. Ley 11.087':
                        cell.alignment = Alignment(horizontal='right')
                        cell.number_format = '0"%"'
                    else:
                        cell.alignment = Alignment(horizontal='right')
                        # Formato de número con separador de miles
                        cell.number_format = '#,##0.00'
            
            # Agregar fila de totales
            totals_row = header_row + len(df) + 1
            worksheet.cell(row=totals_row, column=1, value='TOTAL')
            worksheet.cell(row=totals_row, column=1).font = Font(bold=True)
            worksheet.cell(row=totals_row, column=1).fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            
            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=totals_row, column=col_num)
                cell.border = self._get_border()
                cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                cell.font = Font(bold=True)
                
                if column in totals:
                    cell.value = totals[column]
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
            
            # Ajustar anchos de columna
            for column in df.columns:
                col_letter = self._get_column_letter(df.columns.get_loc(column) + 1)
                max_length = max(
                    df[column].astype(str).str.len().max(),
                    len(column)
                )
                worksheet.column_dimensions[col_letter].width = min(max_length + 3, 50)
            
            # Congelar primera fila
            worksheet.freeze_panes = 'A4'
        
        logger.info(f"Excel exportado exitosamente: {output_path}")
    
    def export_to_csv(self, df: pd.DataFrame, output_path: str) -> None:
        """
        Exporta DataFrame a CSV.
        
        Args:
            df: DataFrame con los datos
            output_path: Ruta del archivo de salida
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        logger.info(f"CSV exportado exitosamente: {output_path}")
    
    @staticmethod
    def _get_column_letter(col_num: int) -> str:
        """Convierte número de columna a letra (1 -> A, 27 -> AA)"""
        letter = ''
        while col_num > 0:
            col_num, remainder = divmod(col_num - 1, 26)
            letter = chr(65 + remainder) + letter
        return letter
    
    @staticmethod
    def _get_border() -> Border:
        """Retorna estilo de borde para celdas"""
        thin_border = Side(border_style="thin", color="000000")
        return Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
